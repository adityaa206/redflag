"""
cost/exporters.py — Export CostRollup to XLSX and CSV formats.

XLSX export uses openpyxl.  CSV export uses the csv stdlib module.

Export is BLOCKED if rollup.export_blocked is True (unflagged review items).
Call rollup.review_acknowledged = True to unblock after human review.
"""
from __future__ import annotations

import csv
import io
import os
from datetime import datetime
from typing import Optional


def _fmt_usd(val: float) -> str:
    """Format a USD value with comma separator."""
    return f"${val:,.0f}"


def _capex_opex_label(item) -> str:
    ce = str(getattr(item.capex_opex, "value", item.capex_opex))
    return ce.upper()


def _category_label(item) -> str:
    cat = str(getattr(item.category, "value", item.category))
    return cat.replace("_", " ").title()


def _flags_label(item) -> str:
    flags = [str(getattr(f, "value", f)) for f in item.review_flags]
    return "; ".join(flags) if flags else ""


# ── CSV export ────────────────────────────────────────────────────────────────

def export_rollup_csv(rollup, output_path: Optional[str] = None) -> bytes:
    """
    Export the CostRollup to CSV.

    If output_path is given, writes to file and returns the bytes.
    If output_path is None, returns bytes only (for Streamlit download).

    Raises RuntimeError if export is blocked (unflagged review items).
    """
    if rollup.export_blocked:
        raise RuntimeError(
            "Export blocked: there are flagged items requiring human review. "
            "Acknowledge them in the Cost & Budget tab before exporting."
        )

    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header
    writer.writerow([
        "Title", "Category", "CapEx/OpEx",
        "Cost Low (USD)", "Cost Base (USD)", "Cost High (USD)",
        "Confidence", "Findings Linked", "Review Flags", "Notes",
    ])

    for item in rollup.line_items:
        writer.writerow([
            item.title,
            _category_label(item),
            _capex_opex_label(item),
            f"{item.cost.low:.0f}",
            f"{item.cost.base:.0f}",
            f"{item.cost.high:.0f}",
            str(getattr(item.confidence, "value", item.confidence)).upper(),
            len(item.finding_ids),
            _flags_label(item),
            item.notes or "",
        ])

    # Summary rows
    writer.writerow([])
    writer.writerow(["TOTALS", "", "",
                     f"{rollup.total.low:.0f}",
                     f"{rollup.total.base:.0f}",
                     f"{rollup.total.high:.0f}",
                     "", "", "", ""])
    writer.writerow(["  CapEx", "", "",
                     f"{rollup.capex_total.low:.0f}",
                     f"{rollup.capex_total.base:.0f}",
                     f"{rollup.capex_total.high:.0f}",
                     "", "", "", ""])
    writer.writerow(["  OpEx", "", "",
                     f"{rollup.opex_total.low:.0f}",
                     f"{rollup.opex_total.base:.0f}",
                     f"{rollup.opex_total.high:.0f}",
                     "", "", "", ""])

    data = buf.getvalue().encode("utf-8")

    if output_path:
        with open(output_path, "wb") as f:
            f.write(data)

    return data


# ── XLSX export ───────────────────────────────────────────────────────────────

def export_rollup_xlsx(rollup, output_path: Optional[str] = None) -> bytes:
    """
    Export the CostRollup to XLSX (openpyxl).

    If output_path is given, writes to file and returns the bytes.
    If output_path is None, returns bytes only.

    Raises RuntimeError if export is blocked.
    Raises ImportError if openpyxl is not installed.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for XLSX export: pip install openpyxl"
        ) from e

    if rollup.export_blocked:
        raise RuntimeError(
            "Export blocked: acknowledge flagged items in the Cost & Budget tab first."
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cost Estimate"

    # Colour scheme
    BG_DARK   = "0E1117"
    BG_HEADER = "0D1828"
    BG_TOTAL  = "080D1C"
    FG_WHITE  = "E1E8F0"
    FG_MUTED  = "5F7CA0"
    FG_RED    = "F43F5E"
    FG_GREEN  = "34D399"
    FG_YELLOW = "FBBF24"

    def _fill(hex_color: str):
        return PatternFill("solid", fgColor=hex_color)

    def _font(hex_color: str = FG_WHITE, bold: bool = False, size: int = 10):
        return Font(color=hex_color, bold=bold, size=size)

    # Title row
    ws.merge_cells("A1:J1")
    ws["A1"] = f"RedFlag — Cost & Remediation Estimate"
    ws["A1"].font   = _font(FG_RED, bold=True, size=14)
    ws["A1"].fill   = _fill(BG_DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Target: {rollup.target or 'N/A'}   |   Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font   = _font(FG_MUTED, size=9)
    ws["A2"].fill   = _fill(BG_DARK)
    ws.row_dimensions[2].height = 18

    # Column headers (row 4)
    headers = [
        "Title", "Category", "CapEx/OpEx",
        "Cost Low (USD)", "Cost Base (USD)", "Cost High (USD)",
        "Confidence", "Findings Linked", "Review Flags", "Notes",
    ]
    col_widths = [38, 16, 10, 14, 14, 14, 10, 14, 22, 28]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font      = _font(FG_WHITE, bold=True)
        cell.fill      = _fill(BG_HEADER)
        cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left",
                                   vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[4].height = 20

    # Data rows
    row_num = 5
    for item in rollup.line_items:
        needs_review = item.needs_review
        bg = "0A0E1A" if row_num % 2 == 0 else BG_DARK

        data = [
            item.title,
            _category_label(item),
            _capex_opex_label(item),
            item.cost.low,
            item.cost.base,
            item.cost.high,
            str(getattr(item.confidence, "value", item.confidence)).upper(),
            len(item.finding_ids),
            _flags_label(item),
            item.notes or "",
        ]

        for col_idx, val in enumerate(data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = _fill("12080A" if needs_review else bg)
            cell.font = _font(FG_RED if needs_review else FG_WHITE)
            if col_idx in (4, 5, 6):
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal="right")

        row_num += 1

    # Blank separator
    row_num += 1

    # Totals
    totals = [
        ("TOTAL",         rollup.total.low,       rollup.total.base,       rollup.total.high,       FG_WHITE),
        ("  CapEx",       rollup.capex_total.low,  rollup.capex_total.base,  rollup.capex_total.high,  FG_YELLOW),
        ("  OpEx",        rollup.opex_total.low,   rollup.opex_total.base,   rollup.opex_total.high,   FG_GREEN),
    ]

    for label, low, base, high, color in totals:
        ws.cell(row=row_num, column=1, value=label).font = _font(color, bold=True)
        for cidx, val in zip([4, 5, 6], [low, base, high]):
            cell = ws.cell(row=row_num, column=cidx, value=val)
            cell.number_format = '"$"#,##0'
            cell.font = _font(color, bold=True)
            cell.alignment = Alignment(horizontal="right")
        for cidx in range(1, 11):
            ws.cell(row=row_num, column=cidx).fill = _fill(BG_TOTAL)
        row_num += 1

    # Freeze header
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    if output_path:
        with open(output_path, "wb") as f:
            f.write(data)

    return data
